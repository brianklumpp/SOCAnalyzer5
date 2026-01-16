import logging
import traceback
from io import BytesIO
from typing import List
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.cell import MergedCell
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select

from ..models import Scan, Control, CUEC, SubserviceOrg
from ..gpt_client import gpt_extract
from ..config import GPT_PROMPTS, FRONTEND_URL

logger = logging.getLogger(__name__)

class ExcelExportService:
    """
    Service for exporting SOC analysis data to Excel template.
    Populates 6 tabs with database content and GPT-generated descriptions.
    """
    
    TEMPLATE_PATH = Path(__file__).resolve().parents[3] / 'data' / 'template' / 'SOC Evaluation Template - 2024.xlsx'
    HIGH_CONFIDENCE_THRESHOLD = 0.7
    
    # Expected tab names in template
    TAB_BASIC_PROCEDURES = "1. Basic Procedures"
    TAB_CONTROL_OBJECTIVES = "2. Control Objective"
    TAB_EXCEPTIONS = "3. Exceptions Noted"
    TAB_CUECS = "4. Complementary User Controls"
    TAB_SUBSERVICE_ORGS = "5. Subservice Orgs"
    TAB_CONCLUSION = "7. Conclusion"
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def _safe_write_cell(self, ws, row: int, col: int, value: any) -> None:
        """
        Safely write to a cell, handling merged cells and sanitizing text.
        Skips writing to merged cells that aren't the top-left cell of the range.
        Removes control characters that openpyxl can't write to Excel.
        """
        try:
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                # Skip writing to merged cells that aren't the master cell
                self.logger.debug(f"[EXCEL_EXPORT] Skipping write to merged cell at ({row}, {col})")
                return
            
            # Sanitize string values to remove control characters
            if isinstance(value, str):
                value = self._clean_text_for_excel(value)
            
            cell.value = value
        except AttributeError as e:
            if "MergedCell" in str(e):
                self.logger.debug(f"[EXCEL_EXPORT] Skipping merged cell at ({row}, {col})")
            else:
                raise
    
    async def generate_report(
        self,
        scan_id: int,
        db: AsyncSession,
        current_user = None
    ) -> BytesIO:
        """
        Main orchestrator: queries database, loads template, populates tabs, returns Excel file.
        
        Args:
            scan_id: ID of scan to export
            db: Database session
            current_user: User who initiated the export (optional)
            
        Returns:
            BytesIO containing Excel file
            
        Raises:
            FileNotFoundError: Template not found
            ValueError: Invalid scan_id or missing data
            Exception: Any processing errors
        """
        print(f"=== EXCEL EXPORT START: scan_id={scan_id} ===", flush=True)
        try:
            print(f"=== Inside try block, about to log ===", flush=True)
            self.logger.info(f"[EXCEL_EXPORT] Starting export for scan_id={scan_id}")
            
            # Query all required data
            self.logger.info(f"[EXCEL_EXPORT] Querying database for scan {scan_id}")
            scan = await self._get_scan(scan_id, db)
            controls = await self._get_controls(scan_id, db)
            cuecs = await self._get_cuecs(scan_id, db)
            suborgs = await self._get_subservice_orgs(scan_id, db)
            
            self.logger.info(
                f"[EXCEL_EXPORT] Data loaded: {len(controls)} controls, "
                f"{len(cuecs)} CUECs, {len(suborgs)} subservice orgs"
            )
            
            # Verify template exists
            if not self.TEMPLATE_PATH.exists():
                raise FileNotFoundError(f"Template not found: {self.TEMPLATE_PATH}")
            
            self.logger.info(f"[EXCEL_EXPORT] Loading template from {self.TEMPLATE_PATH}")
            wb = load_workbook(self.TEMPLATE_PATH)
            
            # Verify all expected tabs exist
            self._verify_template_tabs(wb)
            
            # Populate each tab
            self.logger.info("[EXCEL_EXPORT] Populating Basic Procedures tab")
            self._populate_basic_procedures(wb[self.TAB_BASIC_PROCEDURES], scan, controls)
            
            self.logger.info("[EXCEL_EXPORT] Populating Control Objectives tab")
            self._populate_control_objectives(wb[self.TAB_CONTROL_OBJECTIVES], controls)
            
            self.logger.info("[EXCEL_EXPORT] Populating Exceptions tab")
            self._populate_exceptions(wb[self.TAB_EXCEPTIONS], controls)
            
            self.logger.info("[EXCEL_EXPORT] Populating CUECs tab")
            self._populate_cuecs(wb[self.TAB_CUECS], cuecs)
            
            self.logger.info("[EXCEL_EXPORT] Populating Subservice Orgs tab")
            self._populate_subservice_orgs(wb[self.TAB_SUBSERVICE_ORGS], suborgs)
            
            self.logger.info("[EXCEL_EXPORT] Populating Conclusion tab")
            self._populate_conclusion(wb[self.TAB_CONCLUSION], scan, current_user)
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            print(f"=== EXCEL EXPORT COMPLETE: scan_id={scan_id} ===", flush=True)
            self.logger.info(f"[EXCEL_EXPORT] Export completed successfully for scan {scan_id}")
            return output
            
        except Exception as e:
            self.logger.error(
                f"[EXCEL_EXPORT] Export failed for scan {scan_id}: {e}\n{traceback.format_exc()}"
            )
            raise
    
    def _verify_template_tabs(self, wb) -> None:
        """Verify all required tabs exist in template, log warnings for missing tabs."""
        required_tabs = [
            self.TAB_BASIC_PROCEDURES,
            self.TAB_CONTROL_OBJECTIVES,
            self.TAB_EXCEPTIONS,
            self.TAB_CUECS,
            self.TAB_SUBSERVICE_ORGS,
            self.TAB_CONCLUSION
        ]
        
        available_tabs = wb.sheetnames
        missing_tabs = [tab for tab in required_tabs if tab not in available_tabs]
        
        if missing_tabs:
            self.logger.warning(
                f"[EXCEL_EXPORT] Template missing expected tabs: {missing_tabs}. "
                f"Available tabs: {available_tabs}"
            )
            raise ValueError(f"Template missing required tabs: {missing_tabs}")
        
        self.logger.info(f"[EXCEL_EXPORT] All required tabs verified in template")
    
    async def _get_scan(self, scan_id: int, db: AsyncSession) -> Scan:
        """Query scan by ID."""
        result = await db.execute(select(Scan).where(Scan.id == scan_id))
        scan = result.scalar_one_or_none()
        if not scan:
            raise ValueError(f"Scan {scan_id} not found")
        return scan
    
    async def _get_controls(self, scan_id: int, db: AsyncSession) -> List[Control]:
        """Query all controls for scan."""
        result = await db.execute(select(Control).where(Control.scan_id == scan_id))
        return result.scalars().all()
    
    async def _get_cuecs(self, scan_id: int, db: AsyncSession) -> List[CUEC]:
        """Query all CUECs for scan."""
        result = await db.execute(select(CUEC).where(CUEC.scan_id == scan_id))
        return result.scalars().all()
    
    async def _get_subservice_orgs(self, scan_id: int, db: AsyncSession) -> List[SubserviceOrg]:
        """Query all subservice organizations for scan."""
        result = await db.execute(select(SubserviceOrg).where(SubserviceOrg.scan_id == scan_id))
        return result.scalars().all()
    
    def _populate_basic_procedures(self, ws, scan: Scan, controls: List[Control]) -> None:
        """
        Populate Basic Procedures tab:
        - C7-C15: Basic metadata from scan
        - C26-C33: GPT-generated descriptions
        """
        try:
            # Define fill color #FCE4D6 (light orange/peach)
            fill_color = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            
            # Basic metadata (cells C7-C15)
            # C7: Audited Organization
            self._safe_write_cell(ws, 7, 3, scan.company or "")
            if scan.company:
                ws.cell(7, 3).fill = fill_color
            
            # C8: Product
            self._safe_write_cell(ws, 8, 3, scan.product or "")
            if scan.product:
                ws.cell(8, 3).fill = fill_color
            
            # C9: URL to report
            report_url = f"{FRONTEND_URL}/app/report/{scan.id}"
            self._safe_write_cell(ws, 9, 3, report_url)
            ws.cell(9, 3).fill = fill_color
            
            # C10: Coverage Period
            coverage_period = ""
            if scan.coverage_start and scan.coverage_end:
                coverage_period = f"{scan.coverage_start:%Y-%m-%d} to {scan.coverage_end:%Y-%m-%d}"
            self._safe_write_cell(ws, 10, 3, coverage_period)
            if coverage_period:
                ws.cell(10, 3).fill = fill_color
            
            # C11: Blank
            self._safe_write_cell(ws, 11, 3, "")
            
            # C12: Blank
            self._safe_write_cell(ws, 12, 3, "")
            
            # C13: Report Type (e.g., "SOC 1 Type 2", "SOC 2 Type 2")
            report_type = ""
            if scan.report_type:
                # Format as "SOC 1 Type 2" or "SOC 2 Type 2"
                type_name = scan.report_type.value  # "SOC1" or "SOC2"
                if type_name == "SOC1":
                    report_type = "SOC 1 Type 2"
                elif type_name == "SOC2":
                    report_type = "SOC 2 Type 2"
                elif type_name == "COMBINED":
                    report_type = "SOC 1/SOC 2 Type 2"
                else:
                    report_type = f"{type_name} Type 2"
            self._safe_write_cell(ws, 13, 3, report_type)
            if report_type:
                ws.cell(13, 3).fill = fill_color
            
            # C14: Auditor
            self._safe_write_cell(ws, 14, 3, scan.auditor or "")
            if scan.auditor:
                ws.cell(14, 3).fill = fill_color
            
            # C15: Report Date
            report_date = f"{scan.report_date:%Y-%m-%d}" if scan.report_date else ""
            self._safe_write_cell(ws, 15, 3, report_date)
            if report_date:
                ws.cell(15, 3).fill = fill_color
            
            self.logger.info(f"[EXCEL_EXPORT] Basic metadata populated (C7-C15)")
            
            # GPT-generated descriptions (C26-C33)
            self.logger.info(f"[EXCEL_EXPORT] About to call _generate_gpt_descriptions")
            self._generate_gpt_descriptions(ws, scan, controls)
            self.logger.info(f"[EXCEL_EXPORT] Finished _generate_gpt_descriptions")
            
        except Exception as e:
            self.logger.error(f"[EXCEL_EXPORT] Error populating basic procedures: {e}\n{traceback.format_exc()}")
            raise
    
    def _generate_gpt_descriptions(self, ws, scan: Scan, controls: List[Control]) -> None:
        """Generate 8 GPT descriptions for cells C26-C33."""
        
        # Define fill color #FCE4D6 (light orange/peach)
        fill_color = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        # Calculate statistics - only high confidence items
        high_confidence_controls = [c for c in controls if (c.control_confidence or 0) >= self.HIGH_CONFIDENCE_THRESHOLD]
        control_count = len(high_confidence_controls)
        deviation_count = sum(1 for c in high_confidence_controls if c.has_deviation)
        high_confidence_count = control_count  # All counted controls are high confidence
        low_confidence_count = 0  # No low confidence controls in our counts
        
        report_type = scan.report_type.value if scan.report_type else "SOC2"
        company = scan.company or "Unknown"
        product = scan.product or "services"
        coverage_start = f"{scan.coverage_start:%Y-%m-%d}" if scan.coverage_start else "N/A"
        coverage_end = f"{scan.coverage_end:%Y-%m-%d}" if scan.coverage_end else "N/A"
        
        self.logger.info(f"[EXCEL_EXPORT] Starting GPT descriptions for C26-C33. Stats: {control_count} controls, {deviation_count} deviations, {high_confidence_count} high confidence")
        
        # C26: Service description
        c26_value = self._call_gpt_safe(
            'excel_export_service_description',
            extractor_name="excel_export_service_desc",
            report_type=report_type,
            company=company,
            product=product,
            coverage_start=coverage_start,
            coverage_end=coverage_end
        )
        self._safe_write_cell(ws, 26, 3, c26_value)
        if c26_value:
            ws.cell(26, 3).fill = fill_color
        
        # C27: Transaction materiality
        c27_value = self._call_gpt_safe(
            'excel_export_transaction_materiality',
            extractor_name="excel_export_materiality",
            report_type=report_type,
            company=company,
            product=product,
            control_count=control_count
        )
        self._safe_write_cell(ws, 27, 3, c27_value)
        if c27_value:
            ws.cell(27, 3).fill = fill_color
        
        # C28: Interaction with Solidigm
        c28_value = self._call_gpt_safe(
            'excel_export_interaction',
            extractor_name="excel_export_interaction",
            company=company,
            product=product,
            report_type=report_type
        )
        self._safe_write_cell(ws, 28, 3, c28_value)
        if c28_value:
            ws.cell(28, 3).fill = fill_color
        
        # C29: Appropriateness evaluation
        c29_value = self._call_gpt_safe(
            'excel_export_appropriateness',
            extractor_name="excel_export_appropriateness",
            report_type=report_type,
            company=company,
            product=product,
            control_count=control_count,
            deviation_count=deviation_count,
            high_confidence_count=high_confidence_count
        )
        self._safe_write_cell(ws, 29, 3, c29_value)
        if c29_value:
            ws.cell(29, 3).fill = fill_color
        
        # C30: Sufficiency evaluation
        coverage_summary = f"{high_confidence_count}/{control_count} high confidence"
        c30_value = self._call_gpt_safe(
            'excel_export_sufficiency',
            extractor_name="excel_export_sufficiency",
            report_type=report_type,
            company=company,
            product=product,
            control_count=control_count,
            high_confidence_count=high_confidence_count,
            deviation_count=deviation_count,
            coverage_summary=coverage_summary
        )
        self._safe_write_cell(ws, 30, 3, c30_value)
        if c30_value:
            ws.cell(30, 3).fill = fill_color
        
        # C31: Exceptions summary - only high confidence controls
        exceptions_list = "\n".join([
            f"- {c.control_id}: {c.deviation_desc[:100]}"
            for c in high_confidence_controls if c.has_deviation and c.deviation_desc
        ][:10])  # Limit to 10 exceptions
        
        print(f"=== C31 DEBUG: deviation_count={deviation_count}, exceptions_list={exceptions_list[:200] if exceptions_list else 'EMPTY'} ===", flush=True)
        self.logger.info(f"[EXCEL_EXPORT] C31 - Deviation count: {deviation_count}, exceptions_list length: {len(exceptions_list)}")
        
        if deviation_count > 0:
            c31_value = self._call_gpt_safe(
                'excel_export_exceptions_summary',
                extractor_name="excel_export_exceptions",
                report_type=report_type,
                control_count=control_count,
                deviation_count=deviation_count,
                exceptions_list=exceptions_list or "None listed"
            )
            self.logger.info(f"[EXCEL_EXPORT] C31 GPT response length: {len(c31_value) if c31_value else 0}")
            self._safe_write_cell(ws, 31, 3, c31_value)
            if c31_value:
                ws.cell(31, 3).fill = fill_color
        else:
            self._safe_write_cell(ws, 31, 3, "No exceptions noted.")
            ws.cell(31, 3).fill = fill_color
        
        # C32: Management assertion check (standard text)
        self._safe_write_cell(ws, 32, 3, "Management's assertion regarding the design and operating effectiveness of controls has been reviewed.")
        ws.cell(32, 3).fill = fill_color
        
        # C33: Further evaluation items
        missing_data_summary = f"{low_confidence_count} low confidence controls"
        c33_value = self._call_gpt_safe(
            'excel_export_further_evaluation',
            extractor_name="excel_export_further_eval",
            report_type=report_type,
            deviation_count=deviation_count,
            low_confidence_count=low_confidence_count,
            missing_data_summary=missing_data_summary
        )
        self.logger.info(f"[EXCEL_EXPORT] C33 GPT response length: {len(c33_value) if c33_value else 0}")
        self._safe_write_cell(ws, 33, 3, c33_value)
        if c33_value:
            ws.cell(33, 3).fill = fill_color
        
        self.logger.info(f"[EXCEL_EXPORT] GPT descriptions generated for C26-C33")
    
    def _format_page_refs(self, page_refs) -> str:
        """
        Format page refs consistently: extract numbers only, prefix with 'Page '.
        Handles: [], ['39'], ['Page 36'], ['infrastructure', '39'], etc.
        Returns: '' for empty, 'Page 39' or 'Page 36, 39' for multiple.
        """
        if not page_refs:
            return ""
        
        try:
            numbers = []
            
            if isinstance(page_refs, list):
                for item in page_refs:
                    item_str = str(item).strip()
                    # Skip non-numeric strings like 'infrastructure'
                    if item_str.isdigit():
                        numbers.append(item_str)
                    # Extract number from 'Page 36' format
                    elif item_str.lower().startswith('page'):
                        parts = item_str.split()
                        if len(parts) > 1 and parts[1].isdigit():
                            numbers.append(parts[1])
            elif isinstance(page_refs, (int, float)):
                numbers.append(str(int(page_refs)))
            elif isinstance(page_refs, str):
                # Try to extract numbers from string
                import re
                found = re.findall(r'\d+', page_refs)
                numbers.extend(found)
            
            if not numbers:
                return ""
            
            # Return with 'Page ' prefix
            return "Page " + ", ".join(numbers)
            
        except Exception as e:
            self.logger.debug(f"[EXCEL_EXPORT] Error formatting page refs: {e}")
            return ""
    
    def _call_gpt_safe(self, prompt_key: str, extractor_name: str, **kwargs) -> str:
        """
        Call GPT with error handling, return empty string on failure.
        Logs warnings for failures but doesn't raise exceptions.
        """
        try:
            prompt_template = GPT_PROMPTS.get(prompt_key, "")
            if not prompt_template:
                self.logger.warning(f"[EXCEL_EXPORT] Prompt key '{prompt_key}' not found in GPT_PROMPTS")
                return ""
            
            prompt = prompt_template.format(**kwargs)
            response = gpt_extract(prompt, extractor_name)
            
            # Clean response
            response = response.strip()
            self.logger.debug(f"[EXCEL_EXPORT] GPT call '{extractor_name}' succeeded: {len(response)} chars")
            return response
            
        except Exception as e:
            self.logger.warning(
                f"[EXCEL_EXPORT] GPT call '{extractor_name}' failed: {e}. Leaving cell blank."
            )
            return ""
    
    def _populate_control_objectives(self, ws, controls: List[Control]) -> None:
        """
        Populate Control Objectives tab with high-confidence controls.
        Starting row 9, columns A-E:
        - Column A: Page ref
        - Column B: Control ID
        - Column C: Control Description
        - Column D: Test Procedure
        - Column E: Notes (deviation notes + analyst notes)
        """
        try:
            # Use control_confidence instead of final_confidence
            high_conf_controls = [
                c for c in controls
                if (c.control_confidence or 0) >= self.HIGH_CONFIDENCE_THRESHOLD
            ]
            
            self.logger.info(
                f"[EXCEL_EXPORT] Writing {len(high_conf_controls)} high-confidence controls "
                f"(threshold={self.HIGH_CONFIDENCE_THRESHOLD}, field=control_confidence)"
            )
            
            row = 9
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Define fill color #FCE4D6 (light orange/peach)
            fill_color = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            
            # Define word wrap alignment
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            
            for control in high_conf_controls:
                # Column A: Page refs (formatted with 'Page ' prefix)
                page_refs_str = self._format_page_refs(control.control_page_refs)
                
                # Column E: Combine deviation notes and analyst notes
                notes = ""
                if control.has_deviation and control.deviation_desc:
                    notes = f"Deviation Notes:\n{control.deviation_desc}"
                
                if control.analyst_notes:
                    if notes:  # Already has deviation notes
                        notes += "\n\n"
                    notes += f"Analyst Notes:\n{control.analyst_notes}"
                
                # Write cells
                self._safe_write_cell(ws, row, 1, page_refs_str)  # Column A: Page Ref
                self._safe_write_cell(ws, row, 2, control.control_id or "")  # Column B: Control ID
                self._safe_write_cell(ws, row, 3, control.control_desc or "")  # Column C: Control Description
                self._safe_write_cell(ws, row, 4, control.control_test or "")  # Column D: Test Procedure
                self._safe_write_cell(ws, row, 5, notes)  # Column E: Notes
                
                # Apply formatting: borders, fill color, word wrap
                for col in range(1, 6):
                    cell = ws.cell(row, col)
                    if not isinstance(cell, MergedCell):
                        cell.border = thin_border
                        cell.fill = fill_color
                        cell.alignment = wrap_alignment
                
                # Set row height based on content (estimate: ~15 points per line, min 30)
                max_lines = 1
                if control.control_desc:
                    max_lines = max(max_lines, len(control.control_desc) // 100 + 1)
                if control.control_test:
                    max_lines = max(max_lines, len(control.control_test) // 100 + 1)
                if notes:
                    max_lines = max(max_lines, notes.count('\n') + 1)
                ws.row_dimensions[row].height = max(30, min(max_lines * 15, 150))
                
                row += 1
            
            self.logger.info(f"[EXCEL_EXPORT] Control objectives populated: {row-9} rows written")
            
        except Exception as e:
            self.logger.error(f"[EXCEL_EXPORT] Error populating control objectives: {e}\n{traceback.format_exc()}")
            raise
    
    def _populate_exceptions(self, ws, controls: List[Control]) -> None:
        """
        Populate Exceptions tab with controls that have deviations.
        Starting row 9, columns A-H:
        - Column A: Sequential number (1, 2, 3...)
        - Column B: Control ID
        - Column C: Control Description
        - Column D: Test Result/Deviation
        - Column E: Page refs
        - Column F: Management Response
        - Column G: Left blank
        - Column H: 'What it Means' notes
        """
        try:
            # Only show high confidence controls with deviations
            exception_controls = [
                c for c in controls 
                if c.has_deviation and (c.control_confidence or 0) >= self.HIGH_CONFIDENCE_THRESHOLD
            ]
            
            self.logger.info(f"[EXCEL_EXPORT] Writing {len(exception_controls)} controls with deviations")
            
            row = 9
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Define fill color #FCE4D6 (light orange/peach)
            fill_color = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            
            # Define word wrap alignment
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            
            seq_num = 1
            for control in exception_controls:
                # Column A: Sequential number
                self._safe_write_cell(ws, row, 1, seq_num)
                
                # Column B: Control ID
                self._safe_write_cell(ws, row, 2, control.control_id or "")
                
                # Column C: Control Description
                self._safe_write_cell(ws, row, 3, control.control_desc or "")
                
                # Column D: Test Result/Deviation
                self._safe_write_cell(ws, row, 4, control.deviation_desc or "")
                
                # Column E: Page refs (formatted with 'Page ' prefix)
                page_refs_str = self._format_page_refs(control.control_page_refs)
                self._safe_write_cell(ws, row, 5, page_refs_str)
                
                # Column F: Management Response
                self._safe_write_cell(ws, row, 6, control.management_response_text or "")
                
                # Column G: Left blank
                self._safe_write_cell(ws, row, 7, "")
                
                # Column H: 'What it Means' notes (deviation summary)
                self._safe_write_cell(ws, row, 8, control.deviation_summary or "")
                
                # Apply formatting: borders, fill color, word wrap to columns A-H
                for col in range(1, 9):
                    cell = ws.cell(row, col)
                    if not isinstance(cell, MergedCell):
                        cell.border = thin_border
                        cell.fill = fill_color
                        cell.alignment = wrap_alignment
                
                # Set row height based on content (estimate: ~15 points per line, min 30)
                max_lines = 1
                if control.control_desc:
                    max_lines = max(max_lines, len(control.control_desc) // 80 + 1)
                if control.deviation_desc:
                    max_lines = max(max_lines, len(control.deviation_desc) // 80 + 1)
                if control.management_response_text:
                    max_lines = max(max_lines, len(control.management_response_text) // 80 + 1)
                if control.analyst_notes:
                    max_lines = max(max_lines, len(control.analyst_notes) // 80 + 1)
                ws.row_dimensions[row].height = max(30, min(max_lines * 15, 150))
                
                row += 1
                seq_num += 1
            
            self.logger.info(f"[EXCEL_EXPORT] Exceptions populated: {row-9} rows written")
            
        except Exception as e:
            self.logger.error(f"[EXCEL_EXPORT] Error populating exceptions: {e}\n{traceback.format_exc()}")
            raise
    
    def _populate_cuecs(self, ws, cuecs: List[CUEC]) -> None:
        """
        Populate CUECs tab with high-confidence CUECs (≥0.7).
        Starting row 8, columns A-E:
        - Column A: Sequential number (1, 2, 3...)
        - Column B: Control Objective title (1-3 words generated from description)
        - Column C: CUEC Description
        - Column D: Blank
        - Column E: Analyst Notes
        """
        try:
            high_conf_cuecs = [
                c for c in cuecs
                if (c.cuec_confidence or 0) >= self.HIGH_CONFIDENCE_THRESHOLD
            ]
            
            self.logger.info(
                f"[EXCEL_EXPORT] Writing {len(high_conf_cuecs)} high-confidence CUECs "
                f"(threshold={self.HIGH_CONFIDENCE_THRESHOLD})"
            )
            
            row = 8
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Define fill color #FCE4D6 (light orange/peach)
            fill_color = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            
            # Define word wrap alignment
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            
            seq_num = 1
            for cuec in high_conf_cuecs:
                # Column A: Sequential number
                self._safe_write_cell(ws, row, 1, seq_num)
                
                # Column B: Generate short title from description using GPT
                cuec_title = self._generate_cuec_title(cuec.cuec_description)
                self._safe_write_cell(ws, row, 2, cuec_title)
                
                # Column C: CUEC Description
                self._safe_write_cell(ws, row, 3, cuec.cuec_description or "")
                
                # Column D: Blank
                self._safe_write_cell(ws, row, 4, "")
                
                # Column E: Analyst Notes
                self._safe_write_cell(ws, row, 5, cuec.analyst_notes or "")
                
                # Apply formatting: borders, fill color, word wrap to columns A-E
                for col in range(1, 6):
                    cell = ws.cell(row, col)
                    if not isinstance(cell, MergedCell):
                        cell.border = thin_border
                        cell.fill = fill_color
                        cell.alignment = wrap_alignment
                
                # Set row height based on content (estimate: ~15 points per line, min 30)
                max_lines = 1
                if cuec.cuec_description:
                    max_lines = max(max_lines, len(cuec.cuec_description) // 100 + 1)
                if cuec.analyst_notes:
                    max_lines = max(max_lines, len(cuec.analyst_notes) // 100 + 1)
                ws.row_dimensions[row].height = max(30, min(max_lines * 15, 150))
                
                row += 1
                seq_num += 1
            
            self.logger.info(f"[EXCEL_EXPORT] CUECs populated: {row-8} rows written")
            
        except Exception as e:
            self.logger.error(f"[EXCEL_EXPORT] Error populating CUECs: {e}\n{traceback.format_exc()}")
            raise
    
    def _generate_cuec_title(self, description: str) -> str:
        """
        Generate a 1-3 word title for a CUEC description using GPT.
        Falls back to first few words if GPT fails.
        """
        if not description:
            return ""
        
        try:
            prompt = f"""Generate a concise 1-3 word title for this control description. Return ONLY the title, nothing else.

Description: {description[:200]}

Title:"""
            
            title = self._call_gpt_safe(
                'excel_export_cuec_title',
                extractor_name='excel_export_cuec_title',
                description=description[:200]
            )
            
            if title and len(title) > 0:
                return title.strip()
            
        except Exception as e:
            self.logger.debug(f"[EXCEL_EXPORT] GPT title generation failed: {e}")
        
        # Fallback: use first 3 words
        words = description.split()[:3]
        return " ".join(words) if words else ""
    
    def _populate_subservice_orgs(self, ws, suborgs: List[SubserviceOrg]) -> None:
        """
        Populate Subservice Orgs tab with high and low confidence organizations.
        High confidence starts at A9, low confidence at A14.
        Uses row insertion to avoid overwriting template content below.
        """
        try:
            # Filter out orgs without names first
            valid_suborgs = [s for s in suborgs if s.name and s.name.strip()]
            
            high_conf_orgs = [
                s for s in valid_suborgs
                if (s.confidence or 0) >= self.HIGH_CONFIDENCE_THRESHOLD
            ]
            low_conf_orgs = [
                s for s in valid_suborgs
                if (s.confidence or 0) < self.HIGH_CONFIDENCE_THRESHOLD
            ]
            
            self.logger.info(
                f"[EXCEL_EXPORT] Writing {len(high_conf_orgs)} high-confidence and "
                f"{len(low_conf_orgs)} low-confidence subservice orgs"
            )
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Define fill colors
            data_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Light orange
            blue_fill = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")  # Light blue
            
            # Define word wrap alignment
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            
            # LOW CONFIDENCE SECTION FIRST (starting at A14) to maintain cell references
            if low_conf_orgs:
                # Insert rows for low confidence orgs (need len-1 extra rows since row 14 exists)
                if len(low_conf_orgs) > 1:
                    ws.insert_rows(15, len(low_conf_orgs) - 1)
                
                row = 14
                seq_num = 1
                for org in low_conf_orgs:
                    # Column A: Sequential number
                    self._safe_write_cell(ws, row, 1, seq_num)
                    
                    # Column B: Subservice Organization name
                    self._safe_write_cell(ws, row, 2, org.name or "")
                    
                    # Column C: Page ref (formatted with 'Page ' prefix)
                    page_ref = self._format_page_refs(org.third_party_page_ref)
                    self._safe_write_cell(ws, row, 3, page_ref)
                    
                    # Column D: Description
                    self._safe_write_cell(ws, row, 4, org.third_party_description or "")
                    
                    # Column E: Relevance (GPT-generated)
                    relevance = self._generate_subservice_relevance(org.name, org.third_party_description)
                    self._safe_write_cell(ws, row, 5, relevance)
                    
                    # Apply formatting to columns A-E
                    for col in range(1, 6):
                        cell = ws.cell(row, col)
                        if not isinstance(cell, MergedCell):
                            cell.border = thin_border
                            cell.fill = data_fill
                            cell.alignment = wrap_alignment
                    
                    # Set row height
                    max_lines = 1
                    if org.third_party_description:
                        max_lines = max(max_lines, len(org.third_party_description) // 80 + 1)
                    if relevance:
                        max_lines = max(max_lines, len(relevance) // 80 + 1)
                    ws.row_dimensions[row].height = max(30, min(max_lines * 15, 150))
                    
                    row += 1
                    seq_num += 1
            
            # HIGH CONFIDENCE SECTION (starting at A9)
            if high_conf_orgs:
                # Insert rows for high confidence orgs (need len-1 extra rows since row 9 exists)
                if len(high_conf_orgs) > 1:
                    ws.insert_rows(10, len(high_conf_orgs) - 1)
                
                row = 9
                seq_num = 1
                for org in high_conf_orgs:
                    # Column A: Sequential number
                    self._safe_write_cell(ws, row, 1, seq_num)
                    
                    # Column B: Subservice Organization name
                    self._safe_write_cell(ws, row, 2, org.name or "")
                    
                    # Column C: Page ref (formatted with 'Page ' prefix)
                    page_ref = self._format_page_refs(org.third_party_page_ref)
                    self._safe_write_cell(ws, row, 3, page_ref)
                    
                    # Column D: Description
                    self._safe_write_cell(ws, row, 4, org.third_party_description or "")
                    
                    # Column E: Relevance (GPT-generated)
                    relevance = self._generate_subservice_relevance(org.name, org.third_party_description)
                    self._safe_write_cell(ws, row, 5, relevance)
                    
                    # Columns F, G, H: Blank with blue fill
                    for col in range(6, 9):
                        self._safe_write_cell(ws, row, col, "")
                    
                    # Apply formatting to columns A-H
                    for col in range(1, 9):
                        cell = ws.cell(row, col)
                        if not isinstance(cell, MergedCell):
                            cell.border = thin_border
                            cell.alignment = wrap_alignment
                            # Columns A-E get orange fill, F-H get blue fill
                            if col <= 5:
                                cell.fill = data_fill
                            else:
                                cell.fill = blue_fill
                    
                    # Set row height
                    max_lines = 1
                    if org.third_party_description:
                        max_lines = max(max_lines, len(org.third_party_description) // 80 + 1)
                    if relevance:
                        max_lines = max(max_lines, len(relevance) // 80 + 1)
                    ws.row_dimensions[row].height = max(30, min(max_lines * 15, 150))
                    
                    row += 1
                    seq_num += 1
            
            self.logger.info(
                f"[EXCEL_EXPORT] Subservice orgs populated: {len(high_conf_orgs)} high conf, {len(low_conf_orgs)} low conf"
            )
            
        except Exception as e:
            self.logger.error(f"[EXCEL_EXPORT] Error populating subservice orgs: {e}\n{traceback.format_exc()}")
            raise
    
    def _generate_subservice_relevance(self, name: str, description: str) -> str:
        """
        Generate a 1-sentence relevance statement for a subservice org using GPT.
        Falls back to generic statement if GPT fails.
        """
        if not name:
            return ""
        
        try:
            relevance = self._call_gpt_safe(
                'excel_export_subservice_relevance',
                extractor_name='excel_export_subservice_relevance',
                name=name or "the service provider",
                description=description[:200] if description else "third-party services"
            )
            
            if relevance and len(relevance) > 0:
                return relevance.strip()
            
        except Exception as e:
            self.logger.debug(f"[EXCEL_EXPORT] GPT relevance generation failed: {e}")
        
        # Fallback: generic statement
        return f"{name} provides third-party services relevant to the control environment."
    
    def _format_control_ids(self, control_ids) -> str:
        """Format control_ids JSON array as comma-separated string."""
        if not control_ids:
            return ""
        
        try:
            if isinstance(control_ids, list):
                # Extract control_id from dict objects if present
                ids = []
                for item in control_ids:
                    if isinstance(item, dict) and 'control_id' in item:
                        ids.append(item['control_id'])
                    elif isinstance(item, str):
                        ids.append(item)
                return ", ".join(ids[:10])  # Limit to 10 for readability
            elif isinstance(control_ids, str):
                return control_ids
            return str(control_ids)
        except Exception as e:
            self.logger.warning(f"[EXCEL_EXPORT] Error formatting control_ids: {e}")
            return ""
    
    def _populate_conclusion(self, ws, scan: Scan, current_user=None) -> None:
        """
        Populate Conclusion tab with executive summary.
        Writes user metadata to B12 and executive summary to named range 'Exec_Summary' (cell B13).
        """
        try:
            # Write user metadata at B12
            from datetime import datetime
            from zoneinfo import ZoneInfo
            
            # Get current time in PST timezone
            pst = ZoneInfo('America/Los_Angeles')
            now = datetime.now(pst)
            date_str = now.strftime('%B %d, %Y')  # e.g., "January 15, 2026"
            time_str = now.strftime('%I:%M %p')  # e.g., "03:45 PM"
            
            user_email = current_user.email if current_user else "Unknown User"
            metadata_text = f"Conclusion recorded by {user_email} on {date_str} at {time_str} PST"
            
            # Write to B12 with bold formatting
            cell_b12 = ws['B12']
            if not isinstance(cell_b12, MergedCell):
                cell_b12.value = metadata_text
                cell_b12.font = Font(bold=True)
                self.logger.info(f"[EXCEL_EXPORT] User metadata written to B12: {metadata_text}")
            
            # Convert executive summary JSON to plain text
            summary_text = self._format_executive_summary_for_excel(scan.executive_summary)
            
            if not summary_text:
                summary_text = "No executive summary available."
                self.logger.warning(f"[EXCEL_EXPORT] No executive summary found for scan {scan.id}")
            
            # Try to write to named range first
            written = False
            try:
                if 'Exec_Summary' in ws.parent.defined_names:
                    # Get the cell reference from the named range
                    exec_summary_range = ws.parent.defined_names['Exec_Summary']
                    # Named ranges return destinations which we need to parse
                    destinations = list(exec_summary_range.destinations)
                    if destinations:
                        sheet_name, cell_ref = destinations[0]
                        if ws.title == sheet_name or sheet_name in ['', None]:
                            cell = ws[cell_ref]
                            if not isinstance(cell, MergedCell):
                                cell.value = summary_text
                                # Apply word wrap
                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                                self.logger.info(f"[EXCEL_EXPORT] Executive summary written to named range 'Exec_Summary' ({cell_ref})")
                                written = True
            except Exception as e:
                self.logger.debug(f"[EXCEL_EXPORT] Could not use named range: {e}")
            
            # Fallback to B13 if named range didn't work
            if not written:
                try:
                    cell = ws['B13']
                    if not isinstance(cell, MergedCell):
                        cell.value = summary_text
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        self.logger.info(f"[EXCEL_EXPORT] Executive summary written to cell B13")
                        written = True
                except Exception as e:
                    self.logger.debug(f"[EXCEL_EXPORT] Could not write to B13: {e}")
            
            if not written:
                self.logger.warning(
                    f"[EXCEL_EXPORT] Could not find writable location for executive summary"
                )
            
            self.logger.info(
                f"[EXCEL_EXPORT] Executive summary processed: {len(summary_text)} chars"
            )
            
        except Exception as e:
            self.logger.error(f"[EXCEL_EXPORT] Error populating conclusion: {e}\n{traceback.format_exc()}")
            # Don't raise - continue with export even if conclusion fails
            self.logger.warning(f"[EXCEL_EXPORT] Continuing export despite conclusion error")
    
    def _format_executive_summary_for_excel(self, summary_data: any) -> str:
        """
        Convert executive summary JSON to plain text suitable for Excel.
        Handles both JSON and legacy string formats.
        """
        if not summary_data:
            return ""
        
        # If it's already a string, return it (strip any markdown/HTML)
        if isinstance(summary_data, str):
            # Try to parse as JSON first
            text = summary_data.strip()
            if text.startswith('{') and text.endswith('}'):
                try:
                    import json
                    summary_data = json.loads(text)
                except:
                    # Not JSON, treat as plain text
                    return self._clean_text_for_excel(text)
            else:
                return self._clean_text_for_excel(text)
        
        # Handle JSON structure
        if not isinstance(summary_data, dict):
            return str(summary_data)
        
        # Build formatted text from JSON sections
        sections = []
        
        # Helper to convert field to string (handles both strings and lists)
        def to_text(value):
            if isinstance(value, list):
                return "\n".join(f"• {item}" for item in value if item)
            return str(value) if value else ""
        
        if summary_data.get('about_company'):
            sections.append("About the Company\n" + to_text(summary_data['about_company']))
        
        if summary_data.get('sox_objective'):
            sections.append("Objective (SOX Review)\n" + to_text(summary_data['sox_objective']))
        
        if summary_data.get('key_findings'):
            sections.append("Key Findings\n" + to_text(summary_data['key_findings']))
        
        if summary_data.get('areas_of_concern'):
            sections.append("Areas of Concern\n" + to_text(summary_data['areas_of_concern']))
        
        # Handle recommendations (may be dict, string, or list)
        if summary_data.get('recommendations'):
            recs = summary_data['recommendations']
            if isinstance(recs, dict):
                rec_text = []
                if recs.get('immediate'):
                    rec_text.append("Immediate Actions:\n" + to_text(recs['immediate']))
                if recs.get('strategic'):
                    rec_text.append("Strategic Improvements:\n" + to_text(recs['strategic']))
                if rec_text:
                    sections.append("Recommendations\n" + "\n\n".join(rec_text))
            else:
                sections.append("Recommendations\n" + to_text(recs))
        
        if summary_data.get('deviations_summary'):
            sections.append("Deviations Summary\n" + to_text(summary_data['deviations_summary']))
        
        if summary_data.get('sox_recommendations'):
            sections.append("SOX Recommendations\n" + to_text(summary_data['sox_recommendations']))
        
        # Join all sections with double line breaks
        text = "\n\n".join(sections)
        
        return self._clean_text_for_excel(text)
    
    def _clean_text_for_excel(self, text: str) -> str:
        """
        Clean text for Excel display - remove markdown, HTML tags, and control characters.
        """
        if not text:
            return ""
        
        import re
        
        # Remove control characters (ASCII 0-31 except tab, newline, carriage return)
        # These cause openpyxl to fail with "cannot be used in worksheets"
        text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', text)
        
        # Remove markdown bold **text**
        text = text.replace('**', '')
        
        # Remove HTML tags
        text = re.sub(r'<[^>]+>', '', text)
        
        # Remove markdown code blocks
        text = re.sub(r'```[a-z]*\n', '', text)
        text = text.replace('```', '')
        
        # Normalize line breaks
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        
        # Remove excessive blank lines (more than 2 consecutive)
        text = re.sub(r'\n{3,}', '\n\n', text)
        
        return text.strip()
